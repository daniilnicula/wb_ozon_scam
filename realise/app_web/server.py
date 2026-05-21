import sqlite3
import threading
import time
from io import BytesIO
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from parse_prices import default_user_data_dir, list_chrome_profiles, parse_rows, try_open_chrome, build_driver


def detect_market(url: str) -> str | None:
    normalized = (url or "").lower()
    if "wildberries" in normalized:
        return "WB"
    if "ozon" in normalized:
        return "OZON"
    if "yandex" in normalized:
        return "YANDEX"
    return None


BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "data.sqlite"

SETTING_KEYS = ("headless", "disable_media")

app = Flask(__name__)

_state_lock = threading.Lock()


def table_has_column(table: str, column: str) -> bool:
    with db() as conn:
        rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r[1] == column for r in rows)


def get_default_profile_id() -> int | None:
    with db() as conn:
        row = conn.execute("SELECT id FROM profiles ORDER BY id LIMIT 1").fetchone()
    return row[0] if row else None


_state = {
    "running": False,
    "started_at": None,
    "finished_at": None,
    "total": 0,
    "done": 0,
    "ok": 0,
    "err": 0,
    "message": "",
}

_shared_driver_lock = threading.Lock()
_shared_driver = None
_shared_driver_cfg = None


def _is_driver_alive(driver) -> bool:
    try:
        return bool(driver.session_id)
    except Exception:
        return False


def _close_shared_driver() -> None:
    global _shared_driver, _shared_driver_cfg
    with _shared_driver_lock:
        if _shared_driver is not None:
            try:
                _shared_driver.quit()
            except Exception:
                pass
            _shared_driver = None
            _shared_driver_cfg = None


def _get_shared_driver(cfg: dict):
    global _shared_driver, _shared_driver_cfg
    with _shared_driver_lock:
        if _shared_driver is not None and _shared_driver_cfg == cfg and _is_driver_alive(_shared_driver):
            return _shared_driver
        _close_shared_driver()
        _shared_driver = build_driver(
            headless=cfg["headless"],
            disable_media=cfg["disable_media"],
            page_load_strategy="none" if cfg["concurrency"] > 1 else "normal",
        )
        _shared_driver_cfg = cfg.copy()
        return _shared_driver


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    with db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS profiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            )
            """
        )
        conn.execute(
            "INSERT OR IGNORE INTO profiles(name) VALUES ('Основной')"
        )
        row = conn.execute(
            "SELECT id FROM profiles WHERE name = ? LIMIT 1",
            ("Основной",),
        ).fetchone()
        default_profile_id = row[0] if row else 1

        items_sql = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='items'"
        ).fetchone()
        if items_sql and "CHECK(market IN ('WB','OZON'))" in items_sql["sql"]:
            conn.execute("ALTER TABLE items RENAME TO items_old")
            conn.execute(
                """
                CREATE TABLE items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    market TEXT NOT NULL CHECK(market IN ('WB','OZON','YANDEX')),
                    url TEXT NOT NULL,
                    title TEXT DEFAULT '',
                    seller TEXT DEFAULT '',
                    price TEXT DEFAULT '',
                    price_card TEXT DEFAULT '',
                    updated_at TEXT DEFAULT '',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    profile_id INTEGER NOT NULL DEFAULT 1,
                    FOREIGN KEY(profile_id) REFERENCES profiles(id)
                )
                """
            )
            old_columns = [r[1] for r in conn.execute("PRAGMA table_info(items_old)").fetchall()]
            if "profile_id" in old_columns:
                conn.execute(
                    "INSERT INTO items (id, market, url, title, seller, price, price_card, updated_at, created_at, profile_id)"
                    " SELECT id, market, url, title, seller, price, price_card, updated_at, created_at, profile_id FROM items_old"
                )
            else:
                conn.execute(
                    "INSERT INTO items (id, market, url, title, seller, price, price_card, updated_at, created_at, profile_id)"
                    " SELECT id, market, url, title, seller, price, price_card, updated_at, created_at, ? FROM items_old",
                    (default_profile_id,),
                )
            conn.execute("DROP TABLE items_old")
        else:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    market TEXT NOT NULL CHECK(market IN ('WB','OZON','YANDEX')),
                    url TEXT NOT NULL,
                    title TEXT DEFAULT '',
                    seller TEXT DEFAULT '',
                    price TEXT DEFAULT '',
                    price_card TEXT DEFAULT '',
                    updated_at TEXT DEFAULT '',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    profile_id INTEGER NOT NULL DEFAULT 1,
                    FOREIGN KEY(profile_id) REFERENCES profiles(id)
                )
                """
            )
        if not table_has_column("items", "seller"):
            conn.execute("ALTER TABLE items ADD COLUMN seller TEXT DEFAULT ''")
        if not table_has_column("items", "profile_id"):
            conn.execute(
                f"ALTER TABLE items ADD COLUMN profile_id INTEGER NOT NULL DEFAULT {default_profile_id}"
            )
            conn.execute(
                "UPDATE items SET profile_id = ? WHERE profile_id IS NULL",
                (default_profile_id,),
            )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS error_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                source TEXT NOT NULL DEFAULT '',
                message TEXT NOT NULL DEFAULT ''
            )
            """
        )


def log_error(source: str, message: str) -> None:
    with db() as conn:
        conn.execute(
            "INSERT INTO error_log(ts, source, message) VALUES (?, ?, ?)",
            (time.strftime("%Y-%m-%d %H:%M:%S"), source, message),
        )


def recent_errors(limit: int = 20) -> list[dict]:
    with db() as conn:
        rows = conn.execute(
            "SELECT id, ts, source, message FROM error_log ORDER BY id DESC LIMIT ?",
            (limit,),
        ).fetchall()
    return [dict(r) for r in rows]


def clear_errors() -> int:
    with db() as conn:
        cur = conn.execute("DELETE FROM error_log")
        return cur.rowcount or 0


def get_setting(key: str, default: str = "") -> str:
    with db() as conn:
        row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with db() as conn:
        conn.execute(
            "INSERT INTO settings(key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (key, value),
        )


def effective_chrome_config() -> dict:
    headless_raw = get_setting("headless", "1")
    headless = headless_raw not in ("0", "false", "False", "")
    disable_media_raw = get_setting("disable_media", "0")
    disable_media = disable_media_raw not in ("0", "false", "False", "")
    keep_open_raw = get_setting("keep_browser_open", "0")
    keep_browser_open = keep_open_raw not in ("0", "false", "False", "")
    concurrency_raw = get_setting("concurrency", "1")
    try:
        concurrency = int(concurrency_raw)
    except ValueError:
        concurrency = 1
    concurrency = max(1, min(20, concurrency))
    return {
        "headless": headless,
        "disable_media": disable_media,
        "keep_browser_open": keep_browser_open,
        "concurrency": concurrency,
    }


@app.route("/")
def index():
    profile_id = request.args.get("profile_id")
    selected_profile_id = None
    with db() as conn:
        profiles = [dict(r) for r in conn.execute("SELECT id, name FROM profiles ORDER BY id").fetchall()]
        if profiles:
            default_profile = profiles[0]["id"]
        else:
            default_profile = None
        if profile_id is not None and profile_id != "":
            try:
                profile_id = int(profile_id)
            except ValueError:
                profile_id = None
        if profile_id is None:
            selected_profile_id = default_profile
        else:
            selected_profile_id = profile_id
        if selected_profile_id is not None:
            rows = conn.execute(
                "SELECT * FROM items WHERE profile_id = ? ORDER BY id",
                (selected_profile_id,),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM items ORDER BY id").fetchall()

    # Найти min и max цены для подсвечивания
    min_price = None
    max_price = None
    prices = []
    for row in rows:
        price_str = (row["price"] or "").strip()
        if price_str and price_str.lower() != "ошибка":
            try:
                p = int(price_str)
                prices.append(p)
            except ValueError:
                pass
    if prices:
        min_price = min(prices)
        max_price = max(prices)

    return render_template(
        "index.html",
        items=rows,
        profiles=profiles,
        current_profile_id=selected_profile_id,
        min_price=min_price,
        max_price=max_price,
    )


@app.post("/add")
def add():
    url = (request.form.get("url") or "").strip()
    profile_id = request.form.get("profile_id", type=int)
    if not url:
        return redirect(url_for("index"))
    market = detect_market(url)
    if not market:
        return redirect(url_for("index"))
    if profile_id is None:
        profile_id = get_default_profile_id()
    with db() as conn:
        conn.execute(
            "INSERT INTO items(market, url, profile_id) VALUES (?, ?, ?)",
            (market, url, profile_id),
        )
    return redirect(url_for("index", profile_id=profile_id))


@app.post("/delete/<int:item_id>")
def delete(item_id: int):
    profile_id = request.form.get("profile_id", type=int)
    if profile_id is None:
        profile_id = get_default_profile_id()
    with db() as conn:
        conn.execute("DELETE FROM items WHERE id = ?", (item_id,))
    return redirect(url_for("index", profile_id=profile_id))


@app.post("/clear_items")
def clear_items():
    profile_id = request.form.get("profile_id", type=int)
    if profile_id is None:
        profile_id = get_default_profile_id()
    with db() as conn:
        conn.execute("DELETE FROM items WHERE profile_id = ?", (profile_id,))
    return redirect(url_for("index", profile_id=profile_id))


@app.post("/update/<int:item_id>")
def update(item_id: int):
    market = (request.form.get("market") or "").strip().upper()
    url = (request.form.get("url") or "").strip()
    if market not in {"WB", "OZON"} or not url:
        return redirect(url_for("index"))
    with db() as conn:
        conn.execute(
            "UPDATE items SET market = ?, url = ? WHERE id = ?",
            (market, url, item_id),
        )
    return redirect(url_for("index"))


def _refresh_worker(item_ids: list[int] | None, profile_id: int | None = None):
    try:
        with db() as conn:
            if item_ids:
                placeholders = ",".join("?" for _ in item_ids)
                cur = conn.execute(
                    f"SELECT id, market, url FROM items WHERE id IN ({placeholders}) ORDER BY id",
                    item_ids,
                )
            else:
                query = "SELECT id, market, url FROM items ORDER BY id"
                params = ()
                if profile_id is not None:
                    query = "SELECT id, market, url FROM items WHERE profile_id = ? ORDER BY id"
                    params = (profile_id,)
                cur = conn.execute(query, params)
            items = [dict(r) for r in cur.fetchall()]

        with _state_lock:
            _state["total"] = len(items)
            _state["done"] = 0
            _state["ok"] = 0
            _state["err"] = 0
            _state["message"] = ""

        if not items:
            return

        rows_for_parser = [
            {"row": it["id"], "market": it["market"], "url": it["url"]}
            for it in items
        ]
        cfg = effective_chrome_config()
        if cfg["keep_browser_open"]:
            driver = _get_shared_driver(cfg)
            results = parse_rows(
                rows_for_parser,
                headless=cfg["headless"],
                disable_media=cfg["disable_media"],
                concurrency=cfg["concurrency"],
                driver=driver,
                dispose_driver=False,
            )
        else:
            _close_shared_driver()
            results = parse_rows(
                rows_for_parser,
                headless=cfg["headless"],
                disable_media=cfg["disable_media"],
                concurrency=cfg["concurrency"],
            )

        now = time.strftime("%Y-%m-%d %H:%M:%S")
        ok = 0
        err = 0
        with db() as conn:
            for res in results:
                item_id = res["row"]
                title = res["title"]
                seller = res.get("seller", "") or ""
                price = res["price"]
                price_card = res["price_card"] or ""
                conn.execute(
                    "UPDATE items SET title=?, seller=?, price=?, price_card=?, updated_at=? WHERE id=?",
                    (title, seller, price, price_card, now, item_id),
                )
                if str(price).lower() == "ошибка" or str(title).lower() == "ошибка":
                    err += 1
                else:
                    ok += 1
                with _state_lock:
                    _state["done"] += 1
                    _state["ok"] = ok
                    _state["err"] = err
    except Exception as e:
        msg = f"{type(e).__name__}: {e}"
        with _state_lock:
            _state["message"] = f"Ошибка: {msg}"
        try:
            log_error("refresh", msg)
        except Exception:
            pass
    finally:
        with _state_lock:
            _state["running"] = False
            _state["finished_at"] = time.strftime("%Y-%m-%d %H:%M:%S")


@app.post("/refresh")
def refresh():
    with _state_lock:
        if _state["running"]:
            return jsonify({"ok": False, "message": "Уже идёт парсинг"}), 409
        _state["running"] = True
        _state["started_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
        _state["finished_at"] = None
        _state["total"] = 0
        _state["done"] = 0
        _state["ok"] = 0
        _state["err"] = 0
        _state["message"] = ""

    raw_ids = request.form.get("ids", "")
    ids = [int(x) for x in raw_ids.split(",") if x.strip().isdigit()] or None
    profile_id = request.form.get("profile_id", type=int)

    t = threading.Thread(target=_refresh_worker, args=(ids, profile_id), daemon=True)
    t.start()
    return jsonify({"ok": True})


@app.get("/status")
def status():
    with _state_lock:
        return jsonify(dict(_state))


@app.get("/items.json")
def items_json():
    with db() as conn:
        rows = conn.execute("SELECT * FROM items ORDER BY id").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/settings", methods=["GET", "POST"])
def settings():
    if request.method == "POST":
        headless = "1" if request.form.get("headless") else "0"
        disable_media = "1" if request.form.get("disable_media") else "0"
        keep_browser_open = "1" if request.form.get("keep_browser_open") else "0"
        concurrency = request.form.get("concurrency", "1").strip()
        try:
            concurrency_value = max(1, min(20, int(concurrency)))
        except ValueError:
            concurrency_value = 1
        set_setting("headless", headless)
        set_setting("disable_media", disable_media)
        set_setting("keep_browser_open", keep_browser_open)
        set_setting("concurrency", str(concurrency_value))
        return redirect(url_for("settings"))

    cfg = effective_chrome_config()
    with db() as conn:
        profiles = [dict(r) for r in conn.execute("SELECT id, name FROM profiles ORDER BY id").fetchall()]
    return render_template(
        "settings.html",
        cfg=cfg,
        profiles=profiles,
        errors=recent_errors(20),
    )


@app.post("/add_profile")
def add_profile():
    name = (request.form.get("profile_name") or "").strip()
    if name:
        with db() as conn:
            conn.execute("INSERT OR IGNORE INTO profiles(name) VALUES (?)", (name,))
    return redirect(url_for("settings"))


@app.post("/delete_profile/<int:profile_id>")
def delete_profile(profile_id: int):
    with db() as conn:
        # Не удалять "Основной" профиль
        row = conn.execute("SELECT name FROM profiles WHERE id = ?", (profile_id,)).fetchone()
        if row and row["name"] != "Основной":
            conn.execute("DELETE FROM items WHERE profile_id = ?", (profile_id,))
            conn.execute("DELETE FROM profiles WHERE id = ?", (profile_id,))
    return redirect(url_for("settings"))


@app.post("/test_chrome")
def test_chrome():
    cfg = effective_chrome_config()
    ok, message = try_open_chrome(headless=cfg["headless"], disable_media=cfg["disable_media"])
    if not ok:
        try:
            log_error("test_chrome", message)
        except Exception:
            pass
    return jsonify({"ok": ok, "message": message})


@app.post("/clear_errors")
def clear_errors_endpoint():
    n = clear_errors()
    return jsonify({"ok": True, "cleared": n})


@app.get("/chrome_profiles.json")
def chrome_profiles_json():
    raw = request.args.get("path", "")
    path = Path(raw) if raw else default_user_data_dir()
    if not path or not path.exists():
        return jsonify({"ok": False, "profiles": [], "message": "Папка не существует"}), 200
    return jsonify({"ok": True, "profiles": list_chrome_profiles(path), "path": str(path)})


@app.post("/import")
def import_txt():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "message": "Файл не выбран"}), 400

    profile_id = request.form.get("profile_id", type=int)
    if profile_id is None:
        profile_id = get_default_profile_id()

    raw = f.read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return jsonify({"ok": False, "message": "Не удалось распознать кодировку файла"}), 400

    added = 0
    skipped = 0
    with db() as conn:
        for line in text.splitlines():
            url = line.strip()
            if not url or not url.lower().startswith(("http://", "https://")):
                if url:
                    skipped += 1
                continue
            market = detect_market(url)
            if not market:
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO items(market, url, profile_id) VALUES (?, ?, ?)",
                (market, url, profile_id),
            )
            added += 1

    return jsonify({"ok": True, "added": added, "skipped": skipped})


def _sanitize_filename(name: str) -> str:
    safe_name = "".join(
        c if c.isalnum() or c in "-_ " else "_" for c in (name or "")
    ).strip()
    safe_name = safe_name.replace(" ", "_")
    return safe_name or "profile"


@app.get("/export_txt")
def export_txt():
    profile_id = request.args.get("profile_id", type=int)
    with db() as conn:
        if profile_id:
            profile = conn.execute(
                "SELECT id, name FROM profiles WHERE id = ?",
                (profile_id,),
            ).fetchone()
            if not profile:
                return redirect(url_for("settings"))
            profile_name = profile["name"]
            rows = conn.execute(
                "SELECT url FROM items WHERE profile_id = ? ORDER BY id",
                (profile_id,),
            ).fetchall()
        else:
            profile_name = "all_profiles"
            rows = conn.execute("SELECT url FROM items ORDER BY id").fetchall()

    text = "\n".join(r["url"] for r in rows)
    buf = BytesIO(text.encode("utf-8"))
    buf.seek(0)

    fname = f"{_sanitize_filename(profile_name)}_{time.strftime('%Y%m%d_%H%M')}.txt"
    return send_file(
        buf,
        mimetype="text/plain; charset=utf-8",
        as_attachment=True,
        download_name=fname,
    )


def _to_num_or_text(s):
    s = (s or "").strip()
    if s.isdigit():
        return int(s)
    return s


@app.get("/export.xlsx")
def export_xlsx():
    profile_id = request.args.get("profile_id", type=int)
    with db() as conn:
        if profile_id:
            rows = conn.execute(
                "SELECT market, seller, title, price, price_card, url, updated_at FROM items WHERE profile_id = ? ORDER BY id",
                (profile_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT market, seller, title, price, price_card, url, updated_at FROM items ORDER BY id"
            ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Цены"
    headers = ["№", "Маркет", "Продавец", "Название", "Цена", "Цена по карте", "Ссылка", "Обновлено"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r["market"],
            r["seller"] or "",
            r["title"] or "",
            _to_num_or_text(r["price"]),
            _to_num_or_text(r["price_card"]),
            r["url"],
            r["updated_at"] or "",
        ])

    widths = [5, 8, 20, 60, 12, 15, 80, 20]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"prices_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


if __name__ == "__main__":
    init_db()
    app.run(host="127.0.0.1", port=5000, debug=False)

